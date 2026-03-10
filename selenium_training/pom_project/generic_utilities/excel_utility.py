from openpyxl import load_workbook
import os

def read_excel():

    path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "files_", "demo_testdata.xlsx")

    workbook = load_workbook(path)

    sheet = workbook.active

    data = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        value = row[1]
        data[key] = value

    return data